"""
Seed script — populates the PostgreSQL database from the XLSX schedule files.

Run from inside the backend container:
    python scripts/seed.py

Or locally (set DATABASE_URL env var first):
    DATABASE_URL=postgresql://transit_user:transit_pass@localhost:5432/transit python scripts/seed.py

Data sources:
  /app/data/bus_monastir_sousse.xlsx  — Bus: Monastir→Sousse
  /app/data/bus_sousse_monastir.xlsx  — Bus: Sousse→Monastir
  /app/data/train_monastir_sousse.xlsx — SNCFT Métro du Sahel: Monastir→Sousse
  /app/data/train_sousse_monastir.xlsx — SNCFT Métro du Sahel: Sousse→Monastir
"""
import sys
import os

# Allow running from repo root too
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from sqlalchemy.orm import Session

from app.database import engine, Base
from app import models

# ── Coordinates for all known stops ───────────────────────────────────────────
# Sourced from: Wikipedia, OpenStreetMap, SNCFT timetable Plus Codes, and web searches.

TRAIN_STOPS_MONASTIR_SOUSSE = [
    # name                  lat        lng
    ("Monastir",            35.7709,   10.8264),
    ("L'Aéroport",          35.7581,   10.7547),
    ("Sahline Sokha",       35.7575,   10.7167),
    ("Sousse Z.Ind",        35.7821,   10.6686),
    ("Sousse Sud",          35.8002,   10.6499),
    ("Sousse Med.V",        35.8254,   10.6370),
    ("Sousse Bab Jedid",    35.8229,   10.6415),
]

# Bus stops Monastir→Sousse (10-column schedule, rows are trips)
# Columns: Monastir | Beb Djedid Monastir | Essalem | Khnis |
#          Bekalta? | Hadjeb El Ayoun? | Hr Ksiba? | Msaken |
#          Sousse Kalaa Kebira | Sousse Bab Jedid
BUS_STOPS_MONASTIR_SOUSSE = [
    # name                           lat       lng
    ("Monastir Bus Station",         35.7738,  10.8258),
    ("Monastir Bab Jedid",           35.7700,  10.8240),
    ("Essalem",                      35.7680,  10.8100),
    ("Khnis",                        35.7650,  10.7900),
    ("El Abassia",                   35.7640,  10.7750),
    ("Msaken Junction",              35.7700,  10.7400),
    ("Hr Ksiba",                     35.7750,  10.7200),
    ("Msaken",                       35.7780,  10.7000),
    ("Sousse Kalaa Kebira",          35.7900,  10.6800),
    ("Sousse Bus Terminal",          35.8144,  10.6307),
]

# Bus stops Sousse→Monastir (10-column schedule, mirror order)
BUS_STOPS_SOUSSE_MONASTIR = list(reversed(BUS_STOPS_MONASTIR_SOUSSE))


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalise_time(val) -> str | None:
    """Clean an Excel cell value into HH:MM string or None."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "—", "-", "None"):
        return None
    # Already HH:MM
    if ":" in s:
        parts = s.split(":")
        h, m = int(parts[0]), int(parts[1])
        return f"{h:02d}:{m:02d}"
    return None


def get_or_create_stop(db: Session, name: str, lat: float, lng: float, stype: str) -> models.Stop:
    stop = db.query(models.Stop).filter_by(name=name, type=stype).first()
    if not stop:
        stop = models.Stop(name=name, lat=lat, lng=lng, type=stype)
        db.add(stop)
        db.flush()
    return stop


def seed_bus(db: Session, filepath: str, direction: str, stops_meta: list):
    """
    Parse a bus XLSX (no header row — each row = one trip, each column = one stop).
    """
    print(f"  Seeding bus {direction} from {filepath} ...")
    line = db.query(models.TransitLine).filter_by(name=f"Bus {direction}", type="bus").first()
    if not line:
        line = models.TransitLine(name=f"Bus {direction}", type="bus", direction=direction)
        db.add(line)
        db.flush()

    # Ensure stops exist
    stop_objs = []
    for (name, lat, lng) in stops_meta:
        s = get_or_create_stop(db, name, lat, lng, "bus")
        stop_objs.append(s)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    trip_count = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        times = [normalise_time(cell) for cell in row]
        # Skip empty rows
        if all(t is None for t in times):
            continue

        trip = models.Trip(line_id=line.id, trip_ref=f"bus-row-{row_idx}")
        db.add(trip)
        db.flush()

        for seq, (stop_obj, t) in enumerate(zip(stop_objs, times)):
            if t is None:
                continue  # bus skips this stop on this trip
            st = models.StopTime(
                trip_id=trip.id,
                stop_id=stop_obj.id,
                sequence=seq,
                departure_time=t,
                arrival_time=t,
            )
            db.add(st)

        trip_count += 1

    wb.close()
    print(f"    → {trip_count} bus trips seeded.")


def seed_train(db: Session, filepath: str, direction: str, stops_meta: list):
    """
    Parse a train XLSX (has header rows, rows 5+ = trips, columns 2+ = stop times).
    Row 1 = header with stop names, Row 2 = Plus Codes (ignored), Row 0 = title.
    """
    print(f"  Seeding train {direction} from {filepath} ...")
    line = db.query(models.TransitLine).filter_by(name=f"SNCFT {direction}", type="train").first()
    if not line:
        line = models.TransitLine(name=f"SNCFT {direction}", type="train", direction=direction)
        db.add(line)
        db.flush()

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    # Row index 1 has: [Train N°, Durée, stop1_name, stop2_name, ...]
    header = rows[1]
    stop_names_in_file = [str(h).strip() if h else "" for h in header[2:]]

    # Build stop objects matching file column order to our metadata
    stop_objs_ordered = []
    for col_name in stop_names_in_file:
        matched = None
        for (name, lat, lng) in stops_meta:
            if name.lower() in col_name.lower() or col_name.lower() in name.lower():
                matched = get_or_create_stop(db, name, lat, lng, "train")
                break
        if matched is None and col_name:
            # Fallback: create with approximate coordinates
            matched = get_or_create_stop(db, col_name, 35.80, 10.72, "train")
        stop_objs_ordered.append(matched)

    trip_count = 0
    # Data starts at row index 5 (0-based) based on XLSX structure
    DATA_START_ROW = 5

    for row in rows[DATA_START_ROW:]:
        if not row or row[0] is None:
            continue
        train_no = str(row[0]).strip()
        if not train_no or train_no.startswith("Source"):
            continue

        times = [normalise_time(cell) for cell in row[2:]]
        if all(t is None for t in times):
            continue

        trip = models.Trip(line_id=line.id, trip_ref=train_no)
        db.add(trip)
        db.flush()

        for seq, (stop_obj, t) in enumerate(zip(stop_objs_ordered, times)):
            if stop_obj is None or t is None:
                continue
            st = models.StopTime(
                trip_id=trip.id,
                stop_id=stop_obj.id,
                sequence=seq,
                departure_time=t,
                arrival_time=t,
            )
            db.add(st)

        trip_count += 1

    wb.close()
    print(f"    → {trip_count} train trips seeded.")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Creating tables...")
    Base.metadata.create_all(bind=engine)

    with Session(engine) as db:
        # Clear existing data
        print("Clearing existing data...")
        db.query(models.StopTime).delete()
        db.query(models.Trip).delete()
        db.query(models.TransitLine).delete()
        db.query(models.Stop).delete()
        db.commit()

        DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")

        print("Seeding bus data...")
        seed_bus(
            db,
            os.path.join(DATA_DIR, "bus_monastir_sousse.xlsx"),
            "monastir_sousse",
            BUS_STOPS_MONASTIR_SOUSSE,
        )
        seed_bus(
            db,
            os.path.join(DATA_DIR, "bus_sousse_monastir.xlsx"),
            "sousse_monastir",
            BUS_STOPS_SOUSSE_MONASTIR,
        )

        print("Seeding train data...")
        seed_train(
            db,
            os.path.join(DATA_DIR, "train_monastir_sousse.xlsx"),
            "monastir_sousse",
            TRAIN_STOPS_MONASTIR_SOUSSE,
        )
        seed_train(
            db,
            os.path.join(DATA_DIR, "train_sousse_monastir.xlsx"),
            "sousse_monastir",
            list(reversed(TRAIN_STOPS_MONASTIR_SOUSSE)),
        )

        db.commit()
        print("✅ Seed complete!")

        # Summary
        n_stops = db.query(models.Stop).count()
        n_trips = db.query(models.Trip).count()
        n_times = db.query(models.StopTime).count()
        print(f"   Stops: {n_stops} | Trips: {n_trips} | Stop-times: {n_times}")


if __name__ == "__main__":
    main()
